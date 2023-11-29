import tkinter as tk
from tkinter import filedialog
import openpyxl

def abrir_excel():
    global workbook, sheet, file_path
    file_path = filedialog.askopenfilename(filetypes=[("Archivos de Excel", "*.xlsx")])
    if file_path:
        cargar_excel()

def cargar_excel():
    global workbook, sheet
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    mostrar_formulario()


def mostrar_formulario():
    for widget in frame2.winfo_children():
        widget.destroy()

    columnas = sheet[1]
    column_names = [cell.value for cell in columnas]

    data = []
    for row in sheet.iter_rows(min_row=2):
        data_row = []
        for cell in row:
            data_row.append(cell.value)
        data.append(data_row)

    for col, column_name in enumerate(column_names):
        label = tk.Label(frame2, text=column_name, width=15, anchor='w')
        label.grid(row=col, column=0, pady=3, sticky='w')  
        entry = tk.Entry(frame2, width=40, justify='left')
        entry.grid(row=col, column=1, sticky='w') 
        if data and data[-1]:
            entry.insert(0, data[-1][col])

#    for col, column_name in enumerate(column_names):
#        label = tk.Label(frame2, text=column_name, width=15)
#        label.grid(row=col, column=0, pady=3)
#        entry = tk.Entry(frame2, width=40)
#        entry.grid(row=col, column=1)
#        if data and data[-1]:
#            entry.insert(0, data[-1][col])



#ESTE PARA MOSTRAR LA PRIMERA FILA DEL EXCEL
#def mostrar_formulario():
#    for widget in frame2.winfo_children():
#        widget.destroy()

#    columnas = sheet[1]
#    column_names = [cell.value for cell in columnas]

#    data = []
#    for row in sheet.iter_rows(min_row=2):
#        data_row = []
#        for cell in row:
#            data_row.append(cell.value)
#        data.append(data_row)

#   for col, column_name in enumerate(column_names):
#        label = tk.Label(frame2, text=column_name, width=15)
#        label.grid(row=col, column=0)
#        entry = tk.Entry(frame2, width=30)
#        entry.grid(row=col, column=1)
#        if data and data[0]:
#            entry.insert(0, data[0][col])

    guardar_button = tk.Button(frame2, text="Limpiar", width=20, bg='#90EE90', font=('Calibri', 12, 'bold'), fg='black', command=agregar_nueva_fila)
    guardar_button.grid(row=len(column_names),column=0, pady=8)

    agregar_fila_button = tk.Button(frame2, text="Guardar", width=20, bg='#90EE90', font=('Calibri', 12, 'bold'), fg='black', command=agregar_nueva_fila)
    agregar_fila_button.grid(row=len(column_names),column=1)
    

#MODIFICAR LA FILA DEL EXCEL QUE MUESTRA
#def guardar_excel():
 #   global workbook, file_path
  #  for col, column_name in enumerate(sheet[1]):
   #     cell = sheet.cell(row=2, column=col + 1)
    #    cell.value = frame2.grid_slaves(row=col, column=1)[0].get()
    #workbook.save(file_path)


def agregar_nueva_fila():
    global sheet
    nueva_fila = [frame2.grid_slaves(row=col, column=1)[0].get() for col in range(len(sheet[1]))]
    sheet.append(nueva_fila)
    workbook.save(file_path)
    mostrar_formulario()

    for widget in frame2.winfo_children():
        if isinstance(widget, tk.Entry):
            widget.delete(0, 'end')

ventana = tk.Tk()
ventana.geometry('500x500')
ventana.title("Editar Excel")

frame1 = tk.Frame(ventana)
frame1.pack()
frame2 = tk.Frame(ventana)
frame2.pack()

boton_abrir = tk.Button(frame1, text="Abrir Excel", width=25, bg='orange',font = ('Calibri',12, 'bold'), fg='black', command=abrir_excel)
boton_abrir.pack(pady=10)

workbook = None
sheet = None
file_path = None

ventana.mainloop()

